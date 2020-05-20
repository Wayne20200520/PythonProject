#!/usr/bin/python
# -*- coding: utf-8 -*-
import csv
import os
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings("ignore")
import openpyxl
import time

#####################################常用方法###################################
def create_sheet():
	wb=openpyxl.load_workbook(r'result.xlsx')
	wb.create_sheet(title='1412',index=0)
	wb.save(r'result.xlsx')

#获取dir路径下的所有文件/文件夹路径
def get_filelist(dir):
		fullnamelist=[]
		for home, dirs, files in os.walk(dir):
				for filenamelist in files:
						fullnamelist.append(os.path.join(home, filenamelist))
		return fullnamelist
#####################################常用方法###################################

if __name__ == "__main__":
	logpath=get_filelist(os.getcwd())
	for i in range(len(logpath)):
		if ".csv" in logpath[i]:
		#	导入需要使用的Pandas库和numpy库，读取并创建数据表，将数据表命名为lc,header：第几行开始 
			print("解析文件路径：",logpath[i])

			#lc=pd.DataFrame(pd.read_csv(logpath[i],encoding='iso-8859-1',header=0))
			lc=pd.DataFrame(pd.read_csv(logpath[i],encoding='gbk',header=0))
			#with open(logpath[i]) as f:
				#print(f)
		#	删除表中全部为NaN的行,axis 就是”轴，数轴“的意思,axis的值对应表示：0轴（行），1轴（列）
			lc=lc.dropna(axis=0, how='all')

			if isinstance(lc["averageCurrent"], object):
				lc["averageCurrent"] = lc["averageCurrent"].astype(str).replace(',','') 
		#	创建用于保存结果的result.xlsx文件
			writer = pd.ExcelWriter('result.xlsx',engine= 'openpyxl')


			mean=lc["averageCurrent"].astype(float).mean()
			print("总样本量电流平均值:"+str(mean)+"\n")
			list1=[["总样本量电流平均值:",mean,'','','','','']]
			#list(lc["issuetype"]) 把issuetype这列变成列表
			all_count=len(list(lc["issuetype"]))
			print("总样本量:"+str(all_count)+"\n")
			ids = list(set(list(lc["issuetype"])))# set：去掉重复项
			print("所有统计到的issuetype:"+str(ids)+"\n")
		#   ascending参数的默认值是True，也就是升序,sort_values(by=['...','...''])
		#	lc.sort_values(by=['TargetCurrent'],ascending=True) #TargetCurrent 所在列按照升序排列
				#print(lc["averageCurrent"])
			for i in range(len(ids)):
			#单列数据筛选并排序，lc[“grade”] == “B”是具体的筛选条件，保留averageCurrent等列，并按averageCurrent升序排列
				ls=lc.loc[(lc["issuetype"] == ids[i]), ["averageCurrent", "imei","issuetype"]]
				#ls.to_csv("1.csv",index = False) #	index = False 把索引去掉
				#ls.to_excel("result.xlsx",sheet_name='1412',index = False)
			#issuetype=ids[i]的样本数
				count=lc.loc[lc["issuetype"] == ids[i]].issuetype.count()
				print("issuetype为"+str(ids[i])+"样本量有"+str(count))
				lg=lc.loc[(lc["issuetype"] == ids[i]) & (lc["averageCurrent"].str.isdigit()), ["averageCurrent","imei","issuetype"]].sort_values(by=['averageCurrent'],ascending=True)
				if ids[i]==0:
					time.sleep(10)
				else:
					time.sleep(2)
				#print("lg:",lg)
				issuetype_mean=lg["averageCurrent"].astype(int).mean()
				print("issuetype:"+str(ids[i])+"电流平均值:"+str(issuetype_mean)+"\n")
				count_more_30=lg.loc[(lg["averageCurrent"].astype(int) >= 30)].averageCurrent.count()
				count_more_60=lg.loc[(lg["averageCurrent"].astype(int) >= 60)].averageCurrent.count()
				print("count_more_30:",count_more_30)
				print("count_more_60:",count_more_60)
				b = "%.2f%%" % ((count_more_30/count)* 100)
				c = "%.2f%%" % ((count_more_30/all_count)* 100)	
				d = "%.2f%%" % ((count_more_60/count)* 100)
				e = "%.2f%%" % ((count_more_60/all_count)* 100)
				print("issuetype:"+str(ids[i])+" "+">=30ma相对占比:"+b+" "+">=30ma绝对占比:"+c)
				print("issuetype:"+str(ids[i])+" "+">=60ma相对占比:"+d+" "+">=60ma绝对占比:"+e)
				print("issuetype:"+str(ids[i])+" "+"解析成功"+"\n")
				list=[[ids[i],count,b,d,c,e,issuetype_mean]]
				#list1=list1.append(list)
				list1=list1+list
				ls.to_excel(excel_writer=writer,sheet_name=str(ids[i]),index = False)
			dateframe=pd.DataFrame(list1,columns=['Issue type','样本数','>=30ma相对占比','>=60ma相对占比','>=30ma绝对占比','>=60ma绝对占比','平均电流'])
			dateframe.to_excel(excel_writer=writer,sheet_name='summary',index = False)
			writer.save()

	






